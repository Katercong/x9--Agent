"""Export human-friendly xlsx mirrors:

  exports/产品总表.xlsx   -- 一行一 SKU
  exports/达人总表.xlsx   -- 一行一达人 + 当前状态摘要
  exports/建联流水.xlsx   -- 一行一事件

Cells in xlsx are read-only mirrors; the SQLite DB is the source of truth.
"""
from __future__ import annotations
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "database.db"
OUT = ROOT / "exports"
OUT.mkdir(exist_ok=True)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="3370FF")
ALIGN_TOP = Alignment(vertical="top", wrap_text=True)


def write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def autofit_widths(ws, max_w: int = 50) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 8
        for c in col:
            v = c.value
            if v is None:
                continue
            s = str(v).split("\n")[0]
            length = sum(2 if ord(ch) > 127 else 1 for ch in s)
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max_w, max_len + 2)


def parse_arr(s):
    if not s:
        return ""
    if isinstance(s, (list, tuple)):
        return "\n".join(str(x) for x in s)
    try:
        v = json.loads(s)
        if isinstance(v, list):
            return "\n".join(str(x) for x in v)
    except Exception:
        pass
    return str(s)


def export_products(con: sqlite3.Connection) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "产品总表"
    headers = [
        "SKU", "类目", "子类", "系列", "英文名", "中文名", "规格", "PCS/包",
        "TikTok", "Temu", "eBay", "独立站", "主推等级", "是否主推",
        "中文卖点", "英文卖点", "痛点", "中文场景", "目标人群(中)",
        "默认佣金率", "匹配等级", "达人画像", "Amazon URL", "短链",
        "TK_Content key", "更新时间",
    ]
    write_header(ws, headers)
    rows = con.execute(
        "SELECT p.*, c.name_zh AS category_name FROM product p "
        "LEFT JOIN category c ON p.category_id=c.id "
        "ORDER BY p.is_main_push DESC, p.id"
    ).fetchall()
    for i, r in enumerate(rows, start=2):
        d = dict(r)
        cells = [
            d.get("sku_code"), d.get("category_name"), d.get("subcategory"),
            d.get("series"), d.get("name_en"), d.get("name_zh"),
            d.get("size_label"), d.get("pcs_per_pack"),
            d.get("price_tiktok"), d.get("price_temu"), d.get("price_ebay"),
            d.get("price_independent"),
            d.get("tier"), "是" if d.get("is_main_push") else "",
            parse_arr(d.get("selling_points_zh")),
            parse_arr(d.get("selling_points_en")),
            parse_arr(d.get("pain_points_zh")),
            parse_arr(d.get("scenarios_zh")),
            d.get("target_audience_zh"),
            d.get("commission_rate_default"),
            parse_arr(d.get("creator_match_levels")),
            d.get("creator_persona_zh"),
            d.get("amazon_url"), d.get("short_url"),
            d.get("tk_content_key"), (d.get("updated_at") or "")[:16],
        ]
        for j, v in enumerate(cells, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = ALIGN_TOP
    autofit_widths(ws, max_w=42)
    wb.save(OUT / "产品总表.xlsx")
    print(f"[export] 产品总表.xlsx  {len(rows)} rows")


def export_creators(con: sqlite3.Connection) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "达人总表"
    headers = [
        "Handle", "平台", "等级", "粉丝数", "粉丝原值", "国家", "30d GMV",
        "平均播放", "PPS", "样品评分", "预估发布率",
        "当前状态", "店铺", "对接人", "首联日期", "最近触达",
        "Email", "WhatsApp", "IG", "主页 URL", "内容标签", "备注",
    ]
    write_header(ws, headers)
    rows = con.execute(
        "SELECT * FROM creator ORDER BY tier, followers DESC NULLS LAST, id"
    ).fetchall()
    for i, r in enumerate(rows, start=2):
        d = dict(r)
        cells = [
            d.get("handle"), d.get("platform"), d.get("tier"),
            d.get("followers"), d.get("followers_raw"), d.get("country"),
            d.get("gmv_30d_usd"), d.get("avg_views"), d.get("pps"),
            d.get("sample_score"), d.get("post_rate_est"),
            d.get("current_status"), d.get("store_assigned"),
            d.get("owner_bd"), d.get("first_contact_date"),
            d.get("last_contact_date"),
            d.get("email"), d.get("whatsapp"), d.get("instagram_handle"),
            d.get("profile_url"),
            parse_arr(d.get("category_tags")), d.get("notes"),
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=i, column=j, value=v).alignment = ALIGN_TOP
    autofit_widths(ws, max_w=40)
    wb.save(OUT / "达人总表.xlsx")
    print(f"[export] 达人总表.xlsx  {len(rows)} rows")


def export_outreach(con: sqlite3.Connection) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "建联流水"
    headers = [
        "日期", "店铺", "对接人", "Handle", "状态", "样品数", "佣金率",
        "视频链接", "授权码", "备注", "创建时间",
    ]
    write_header(ws, headers)
    rows = con.execute(
        "SELECT o.*, c.handle FROM outreach o JOIN creator c ON c.id=o.creator_id "
        "ORDER BY o.event_date DESC, o.id DESC"
    ).fetchall()
    for i, r in enumerate(rows, start=2):
        d = dict(r)
        cells = [
            d.get("event_date"), d.get("store_name"), d.get("bd_owner"),
            d.get("handle"), d.get("status"), d.get("sample_qty"),
            d.get("commission_rate"), d.get("video_url"), d.get("ad_auth_code"),
            d.get("remark"), (d.get("created_at") or "")[:16],
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=i, column=j, value=v).alignment = ALIGN_TOP
    autofit_widths(ws, max_w=50)
    wb.save(OUT / "建联流水.xlsx")
    print(f"[export] 建联流水.xlsx  {len(rows)} rows")


def main() -> None:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    export_products(con)
    export_creators(con)
    export_outreach(con)
    con.close()


if __name__ == "__main__":
    main()
