"""Acceptance tests covering spec section 17."""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timedelta
import pytest

from x9_creator_desktop_system.backend.database import SessionLocal
from x9_creator_desktop_system.backend.models.creator import Creator


def test_health(client):
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_migrate_idempotent(client):
    assert client.post("/api/local/db/migrate").status_code == 200
    assert client.post("/api/local/db/migrate").status_code == 200


def test_extension_heartbeat(client):
    r = client.post("/api/local/extension/heartbeat", json={
        "event_type": "extension_heartbeat",
        "extension_id": "x9_tiktok_collector",
        "extension_version": "3.0.0",
        "worker_id": "test_worker",
        "account_id": "test_account",
        "browser_profile": "chrome_default",
        "current_url": "https://www.tiktok.com/@example",
        "tiktok_page_status": "on_tiktok",
        "tiktok_login_status": "logged_in",
        "page_type": "creator_profile",
        "active_tab_title": "TikTok",
        "timestamp": "2026-05-08T00:00:00Z",
    })
    assert r.status_code == 200
    assert r.json()["status"] == "online"
    s = client.get("/api/local/extension/status").json()
    assert s["any_online"] is True


def test_observation_ingest(client):
    payload = {
        "event_type": "creator_observation",
        "platform": "tiktok",
        "search_keyword": "sanitary pads",
        "creator": {
            "handle": "test_handle", "display_name": "Test",
            "bio": "test", "followers_raw": "12.3K",
            "email": "test@gmail.com", "external_links": [],
            "store_assigned": "X9x9 Shop 01", "owner_bd": "Mercy",
        },
        "source_video": {"video_url": "https://www.tiktok.com/@x/video/1", "title": "t", "description": "d", "hashtags": []},
        "collected_at": "2026-05-08T00:00:00Z",
    }
    r1 = client.post("/api/local/collector/observations", json=payload)
    assert r1.json()["action"] == "inserted"
    assert r1.json()["pipeline"]["ok"] is True
    payload["collected_at"] = "2026-05-08 12:30:00"
    r2 = client.post("/api/local/collector/observations", json=payload)
    assert r2.json()["action"] == "updated"

    db = SessionLocal()
    try:
        c = db.query(Creator).filter_by(handle="test_handle").one()
        assert c.followers_count == 12300
    finally:
        db.close()

    listed = client.get("/api/local/creators?limit=50").json()["items"]
    saved = next(item for item in listed if item["handle"] == "test_handle")
    assert saved["collected_at"].startswith("2026-05-08T12:30:00")
    assert saved["store_assigned"] == "X9x9 Shop 01"
    assert saved["owner_bd"] == "Mercy"

    filtered = client.get("/api/local/creators?collected_date=2026-05-08&limit=50").json()["items"]
    assert any(item["handle"] == "test_handle" for item in filtered)


def test_observation_enrichment_extracts_collected_profile_text(client):
    handle = f"enriched_creator_{datetime.now():%Y%m%d%H%M%S%f}"
    payload = {
        "event_type": "creator_observation",
        "platform": "tiktok",
        "search_keyword": "ugc creator",
        "creator": {},
        "raw_profile": {
            "username": handle,
            "nickname": "Enriched Creator",
            "profile_url": f"https://www.tiktok.com/@{handle}",
            "visible_text": (
                "Enriched Creator\n"
                "18.5K followers\n"
                "Business: enriched.creator@example.com\n"
                "Instagram https://instagram.com/enriched_creator\n"
                "WhatsApp +15551234567"
            ),
            "external_links": [],
        },
        "source_video": {
            "video_url": f"https://www.tiktok.com/@{handle}/video/123",
            "title": "ugc review",
            "description": "home routine",
        },
        "collected_at": "2026-05-18T08:00:00Z",
    }
    r = client.post("/api/local/collector/observations", json=payload)
    assert r.status_code == 200
    assert r.json()["action"] == "inserted"

    db = SessionLocal()
    try:
        saved = db.query(Creator).filter_by(handle=handle).one()
        assert saved.email == "enriched.creator@example.com"
        assert saved.followers_count == 18500
        assert "instagram.com/enriched_creator" in (saved.external_links_json or "")
        snapshot = json.loads(saved.profile_snapshot_json or "{}")
        assert snapshot["emails"] == ["enriched.creator@example.com"]
        assert "visible_text" in " ".join(snapshot["source_text_fields"])
        assert "WhatsApp" in snapshot["text_excerpt"]
    finally:
        db.close()


def test_no_contact_tiktok_observation_is_removed_by_server(client):
    handle = f"no_contact_collect_{datetime.now():%Y%m%d%H%M%S%f}"
    payload = {
        "event_type": "creator_observation",
        "platform": "tiktok",
        "search_keyword": "home routine",
        "creator": {
            "handle": handle,
            "display_name": "No Contact Creator",
            "bio": "home cleaning and family routine",
            "followers_raw": "20K",
            "email": None,
            "external_links": [],
        },
        "source_video": {
            "video_url": f"https://www.tiktok.com/@{handle}/video/456",
            "title": "morning routine",
            "description": "home cleaning",
        },
    }
    r = client.post("/api/local/collector/observations", json=payload)
    assert r.status_code == 200
    body = r.json()
    assert body["action"] == "skipped"
    assert body["reason"] == "missing_contact"
    assert body["creator_id"] is None

    db = SessionLocal()
    try:
        assert db.query(Creator).filter_by(handle=handle).first() is None
    finally:
        db.close()


def test_contact_methods_from_creator_bio_filter(client):
    payload = {
        "event_type": "creator_observation",
        "platform": "tiktok",
        "search_keyword": "creator collab",
        "creator": {
            "handle": "whatsapp_creator",
            "display_name": "WhatsApp Creator",
            "bio": "Collabs via WhatsApp 15551234567 or IG @whatsapp_creator",
            "followers_raw": "18K",
            "email": None,
            "external_links": ["https://external.example/not-used-for-contact-filter"],
        },
        "source_video": {"video_url": "https://www.tiktok.com/@x/video/5", "title": "review", "description": "ugc"},
        "collected_at": "2026-05-08T10:00:00Z",
    }
    assert client.post("/api/local/collector/observations", json=payload).status_code == 200

    listed = client.get("/api/local/creators?contact_channel=whatsapp&limit=50").json()["items"]
    saved = next(item for item in listed if item["handle"] == "whatsapp_creator")
    assert saved["has_contact"] is True
    assert "whatsapp" in saved["contact_types"]
    assert "instagram" in saved["contact_types"]
    assert any(method["type"] == "whatsapp" for method in saved["contact_methods"])
    assert any(method["type"] == "instagram" and method["value"] == "@whatsapp_creator" for method in saved["contact_methods"])

    by_text = client.get("/api/local/creators?contact_contains=15551234567&limit=50").json()["items"]
    assert any(item["handle"] == "whatsapp_creator" for item in by_text)


def test_creator_table_import_csv_runs_pipeline(client):
    csv_body = (
        "handle,platform,display_name,bio,followers,email,whatsapp,instagram_handle,category_tags,source,current_status\n"
        "table_import_creator,tiktok,Table Import Creator,"
        "\"Period care UGC creator\",25000,,15551234567,@table_import_ig,"
        "\"[\"\"feminine_care\"\", \"\"ugc\"\"]\",table_test,已寄样。\n"
    ).encode("utf-8")
    r = client.post(
        "/api/local/import/creators/table?filename=creators.csv",
        content=csv_body,
        headers={"Content-Type": "text/csv"},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["inserted"] == 1
    assert data["failed"] == 0
    assert data["items"][0]["pipeline"]["ok"] is True

    listed = client.get("/api/local/creators?handle_contains=table_import_creator&limit=5").json()["items"]
    saved = next(item for item in listed if item["handle"] == "table_import_creator")
    assert saved["recommendation_status"] is not None
    assert saved["current_status"] == "已寄样"
    assert "whatsapp" in saved["contact_types"]
    assert "instagram" in saved["contact_types"]

    by_status = client.get("/api/local/creators?current_status=已寄样&limit=50").json()["items"]
    assert any(item["handle"] == "table_import_creator" for item in by_status)


def test_creator_table_import_xlsx_runs_pipeline(client):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "handle",
        "platform",
        "display_name",
        "bio",
        "followers",
        "email",
        "whatsapp",
        "instagram_handle",
        "category_tags",
        "source",
    ])
    sheet.append([
        "xlsx_import_creator",
        "tiktok",
        "XLSX Import Creator",
        "Excel UGC creator",
        32100,
        "",
        "+1 555 987 6543",
        "@xlsx_import_ig",
        '["excel", "ugc"]',
        "xlsx_test",
    ])
    out = io.BytesIO()
    workbook.save(out)
    workbook.close()

    r = client.post(
        "/api/local/import/creators/table?filename=creators.xlsx",
        content=out.getvalue(),
        headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["inserted"] == 1
    assert data["failed"] == 0
    assert data["items"][0]["pipeline"]["ok"] is True

    db = SessionLocal()
    try:
        saved = db.query(Creator).filter_by(handle="xlsx_import_creator").one()
        assert saved.source == "table_import"
        assert saved.followers_count == 32100
    finally:
        db.close()


def test_creator_table_import_infers_handle_from_profile_url_alias(client):
    handle = f"url_alias_creator_{datetime.now():%Y%m%d%H%M%S%f}"
    csv_body = (
        "url,nickname,signature,followers,email\n"
        f"https://www.tiktok.com/@{handle},URL Alias Creator,"
        "\"UGC creator with email url_alias@example.com\",12.5K,url_alias@example.com\n"
    ).encode("utf-8")

    r = client.post(
        "/api/local/import/creators/table?filename=tiktok-leads.csv",
        content=csv_body,
        headers={"Content-Type": "text/csv"},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["inserted"] == 1
    assert data["failed"] == 0

    db = SessionLocal()
    try:
        saved = db.query(Creator).filter_by(handle=handle).one()
        assert saved.display_name == "URL Alias Creator"
        assert saved.followers_count == 12500
    finally:
        db.close()


def test_creator_table_import_infers_handle_from_raw_text(client):
    handle = f"raw_text_creator_{datetime.now():%Y%m%d%H%M%S%f}"
    csv_body = (
        "raw_text\n"
        f"\"@{handle}\nRaw Text Creator\n18K followers\nBusiness raw_text@example.com\"\n"
    ).encode("utf-8")

    r = client.post(
        "/api/local/import/creators/table?filename=tiktok_no_more_results_text.csv",
        content=csv_body,
        headers={"Content-Type": "text/csv"},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["inserted"] == 1
    assert data["failed"] == 0

    db = SessionLocal()
    try:
        saved = db.query(Creator).filter_by(handle=handle).one()
        assert saved.profile_url == f"https://www.tiktok.com/@{handle}"
        assert saved.email == "raw_text@example.com"
    finally:
        db.close()


def test_creator_import_xlsx_template_downloads(client):
    from openpyxl import load_workbook

    r = client.get("/api/local/import/creators/template.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "creator-import-template.xlsx" in r.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    finally:
        workbook.close()
    assert headers[:4] == ["handle", "platform", "profile_url", "display_name"]


def test_search_keyword_only_match(client):
    client.post("/api/local/collector/observations", json={
        "event_type": "creator_observation",
        "platform": "tiktok",
        "search_keyword": "sanitary pads",
        "creator": {"handle": "search_only_creator", "bio": "just vibing in NYC",
                    "followers_raw": "30K", "email": "hi@gmail.com", "display_name": "x"},
        "source_video": {"video_url": "https://www.tiktok.com/@x", "title": "TikTok - Make Your Day", "description": None},
    })
    client.post("/api/local/process/run-full-pipeline", json={})
    db = SessionLocal()
    try:
        c = db.query(Creator).filter_by(handle="search_only_creator").one()
        assert "search_keyword_only_match" in (c.risk_tags_json or "")
        assert "manual_review_required" not in (c.risk_tags_json or "")
        assert c.queue_type == "low_confidence_hold"
        assert c.review_required == 0
        assert c.recommendation_status == "hold"
    finally:
        db.close()


def test_recommendation_queues(client):
    samples = [
        # Strong feminine -> conversion queue
        ("fem_strong", "menstrual care reviews and period products. business: hi@me.co",
         "60K", "hi@me.co", "feminine care",
         "https://www.tiktok.com/@x/video/1", "my period routine", "panty liner reviews"),
        # Medium feminine, good DQ -> warm lead
        ("fem_medium", "skincare and self care lifestyle, occasional period care picks",
         "120K", "team@selfcare.com", "feminine care",
         "https://www.tiktok.com/@x/video/2", "morning routine", "self care wellness"),
        # Macro low fit -> brand awareness
        ("macro_low", "comedy creator since 2018", "5M", "agent@bigtalent.com", "vintage",
         "https://www.tiktok.com/@x/video/3", "skit time", "lol"),
    ]
    for handle, bio, fol, email, kw, video, title, desc in samples:
        client.post("/api/local/collector/observations", json={
            "event_type": "creator_observation",
            "platform": "tiktok",
            "search_keyword": kw,
            "creator": {"handle": handle, "bio": bio, "followers_raw": fol,
                        "email": email, "display_name": handle, "external_links": []},
            "source_video": {"video_url": video, "title": title, "description": desc, "hashtags": []},
        })
    db = SessionLocal()
    try:
        db.merge(Creator(
            id="direct_no_email_creator",
            platform="tiktok",
            handle="no_email_creator",
            display_name="no_email_creator",
            bio="mom life and home routine",
            followers_raw="20K",
            followers_count=20000,
            search_keyword="lifestyle",
            source_video_url="https://www.tiktok.com/@x/video/4",
            source_video_title="morning routine for moms",
            source_video_description="home cleaning",
            has_email=0,
        ))
        db.commit()
    finally:
        db.close()
    client.post("/api/local/process/run-full-pipeline", json={})

    expected = {
        "fem_strong": "feminine_conversion_queue",
        "fem_medium": "feminine_warm_lead_queue",
        "macro_low": "macro_brand_awareness_queue",
        "no_email_creator": "no_contact_info_queue",
    }
    db = SessionLocal()
    try:
        for handle, queue in expected.items():
            c = db.query(Creator).filter_by(handle=handle).one()
            assert c.queue_type == queue, f"@{handle}: got {c.queue_type}, expected {queue}"
    finally:
        db.close()


def test_creator_list_sorting(client):
    now = datetime(2026, 5, 8, 12, 0, 0)
    rows = [
        Creator(
            id="sort_case_big",
            platform="tiktok",
            handle="sort_case_big",
            followers_count=900_000,
            email="big@example.com",
            has_email=1,
            outreach_priority="P3",
            recommendation_score=40,
            primary_product_fit_score=35,
            recommendation_status="recommended",
            current_status="待建联",
            collected_at=now - timedelta(days=2),
        ),
        Creator(
            id="sort_case_strong",
            platform="tiktok",
            handle="sort_case_strong",
            followers_count=40_000,
            email="strong@example.com",
            has_email=1,
            outreach_priority="P2",
            recommendation_score=96,
            primary_product_fit_score=92,
            recommendation_status="recommended",
            current_status="待回复",
            collected_at=now - timedelta(days=1),
        ),
        Creator(
            id="sort_case_published",
            platform="tiktok",
            handle="sort_case_published",
            followers_count=80_000,
            email="published@example.com",
            has_email=1,
            outreach_priority="P4",
            recommendation_score=50,
            primary_product_fit_score=45,
            recommendation_status="recommended",
            current_status="视频已发布",
            collected_at=now - timedelta(hours=12),
        ),
        Creator(
            id="sort_case_urgent",
            platform="tiktok",
            handle="sort_case_urgent",
            followers_count=120_000,
            email=None,
            has_email=0,
            outreach_priority="P1",
            recommendation_score=70,
            primary_product_fit_score=65,
            recommendation_status="recommended",
            current_status="已寄样",
            collected_at=now,
        ),
    ]
    db = SessionLocal()
    try:
        for row in rows:
            db.merge(row)
        db.commit()
    finally:
        db.close()

    by_followers = client.get("/api/local/creators?handle_contains=sort_case_&sort_by=followers&limit=10").json()["items"]
    assert by_followers[0]["handle"] == "sort_case_big"

    by_score = client.get("/api/local/creators?handle_contains=sort_case_&sort_by=score&limit=10").json()["items"]
    assert by_score[0]["handle"] == "sort_case_strong"

    by_priority = client.get("/api/local/creators?handle_contains=sort_case_&sort_by=priority&limit=10").json()["items"]
    assert by_priority[0]["handle"] == "sort_case_urgent"

    by_contact = client.get("/api/local/creators?handle_contains=sort_case_&sort_by=contactable&limit=10").json()["items"]
    assert by_contact[0]["has_email"] is True

    follower_window = client.get("/api/local/creators?handle_contains=sort_case_&min_followers=100000&max_followers=200000&limit=10").json()["items"]
    assert [item["handle"] for item in follower_window] == ["sort_case_urgent"]

    current_status = client.get("/api/local/creators?handle_contains=sort_case_&current_status=视频已发布&limit=10").json()["items"]
    assert [item["handle"] for item in current_status] == ["sort_case_published"]

    reply_status = client.get("/api/local/creators?handle_contains=sort_case_&current_status=待回复&limit=10").json()["items"]
    assert [item["handle"] for item in reply_status] == ["sort_case_strong"]

    pending_status = client.get("/api/local/creators?handle_contains=sort_case_&current_status=待建联&limit=10").json()["items"]
    assert [item["handle"] for item in pending_status] == ["sort_case_big"]

    assert client.get("/api/local/creators?sort_by=unknown").status_code == 400


def test_no_manual_review_task_created_for_search_only_hold(client):
    client.post("/api/local/collector/observations", json={
        "event_type": "creator_observation",
        "platform": "tiktok",
        "search_keyword": "sanitary pads",
        "creator": {"handle": "no_review_task_creator", "bio": "daily life",
                    "followers_raw": "25K", "email": "hello@gmail.com", "display_name": "x"},
        "source_video": {"video_url": "https://www.tiktok.com/@x", "title": "TikTok - Make Your Day", "description": None},
    })
    client.post("/api/local/process/run-full-pipeline", json={})
    tasks = client.get("/api/local/review-tasks?status=pending").json()
    assert not any(item["creator_id"].endswith("no_review_task_creator") for item in tasks["items"])


def test_export_csv(client):
    r = client.get("/api/local/export/recommended-creators.csv")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    body = r.content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(body))
    expected = {
        "handle", "display_name", "profile_url", "followers_count", "email",
        "contact_types", "contact_methods",
        "recommended_product_type", "recommended_collab_type", "outreach_priority",
        "current_status", "recommendation_status", "recommendation_reason", "risk_tags",
        "next_action", "notes",
    }
    assert expected.issubset(set(reader.fieldnames or []))
